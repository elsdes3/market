#!/usr/bin/env python3
# -*- coding: utf-8 -*-


"""Define utilities to format XLSX files."""

# pylint: disable=invalid-name,dangerous-default-value
# pylint: disable=too-many-locals,unused-argument,redefined-outer-name


from datetime import datetime

import pandas as pd
import pytz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def convert_colnums_to_colnames(n: int) -> str:
    """Convert spreadsheet column number to alphabetical name."""
    conv_string = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    base = 26
    i = n - 1
    if i < base:
        return conv_string[i]
    return convert_colnums_to_colnames(i // base) + conv_string[i % base]


def export_df_to_formatted_spreadsheet(df: pd.DataFrame, fpath: str) -> None:
    """Export DataFrame to formattedworksheet in XLSX file."""
    start_time = datetime.now(pytz.timezone("US/Eastern"))
    start_time_str = start_time.strftime("%Y-%m-%d %H:%M:%S.%f")
    print(f"Export start time = {start_time_str[:-3]}...", end="")
    wb = Workbook()
    wb.iso_dates = True
    ws = wb.active

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Set column header to bold font and hilight with yellow background
    for cell in ws["1:1"]:
        cell.font = Font(
            color="000000", bold=True, name="Calibri", underline="single"
        )

    # Align cell contents vertically and horizontally
    for row in ws:
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set bottom border for column header row
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    for cell in ws[1]:
        cell.border = Border(bottom=thick, left=thin, right=thin, top=thick)

    # Get list of cell header column names (eg. A, B, etc.)
    xls_col_names = [
        convert_colnums_to_colnames(k + 1) for k in range(len(list(df)))
    ]

    # Iterate over list of cell header column names and column header names
    for k, (col_name, xls_col_name) in enumerate(zip(list(df), xls_col_names)):
        # Get length of column header
        header_col_width = len(col_name)
        # Get maximum length of column values
        values_col_width_max = df[col_name].astype(str).str.len().max()
        # Set column width with a buffer of two whitespaces added to the end
        desired_col_width = max(header_col_width, values_col_width_max) + 2
        # Set column width
        ws.column_dimensions[xls_col_name].width = desired_col_width + 2
        # print(
        #     k+1,
        #     xls_col_name,
        #     col_name,
        #     header_col_width,
        #     values_col_width_max,
        #     desired_col_width,
        # )

    wb.save(fpath)
    end_time = datetime.now(pytz.timezone("US/Eastern"))
    end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S.%f")
    duration = end_time - start_time
    duration = duration.seconds + (duration.microseconds / 1_000_000)
    print(f"done at {end_time_str[:-3]} ({duration:.3f} seconds).")

    wb.close()
